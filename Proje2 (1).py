

import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
program_cikti_s=11
ders_cikti_s=6



program_cikti = list(range(1, (program_cikti_s)))  
ders_cikti = list(range(1, (ders_cikti_s)))  

iliskiler = [
    [1, 1, 0, 1, 1],
    [0, 0, 1, 0.2, 0],
    [0, 0, 0.5 , 1, 0],
    [0, 0, 0, 0.8, 1],
    [0, 1, 0, 1, 0],
    [1, 0, 0, 0, 1],
    [0, 0, 0, 1, 1],
    [0, 1, 1, 1, 1],
    [1, 0, 1, 1, 0],
    [0, 1, 1, 1, 0],
]




df = pd.DataFrame(iliskiler, columns=ders_cikti, index=program_cikti)


iliskiler_deg = []

katsayi = 1 / (ders_cikti_s - 1)

for index, row in df.iterrows():
    toplam = 0
    for value in row:  
        toplam += value * katsayi
    iliskiler_deg.append(round(toplam, 2))  


df["İlişki Değ."] = iliskiler_deg

file_name = "ders_program_cikti_iliskisi.xlsx"
if os.path.exists(file_name):
    print(f"{file_name} dosyası zaten mevcut. Yeni dosya oluşturulmadı.")
else:
    df.to_excel(file_name, sheet_name="Tablo 1", index_label="Prg Çıktı")
    print(f"{file_name} başarıyla oluşturuldu.")



ders_cikti_s = 5  
degerlendirme_sayisi_s = 6  


degerlendirme_sutunlari = ["Ödev1", "Ödev2", "Quiz", "Quiz4", "Vize", "Final"]
agirliklar = [10, 10, 10, 10, 20, 40]  

iliskiler_ders = [
     [1, 1, 0, 1, 1, 1],
    [0, 1, 1, 1, 0, 1],
    [1, 0, 1, 0, 1, 0],
    [0, 1, 0, 1, 1, 1],
    [1, 0, 1, 0, 0, 1]
]


df_degerlendirme = pd.DataFrame(iliskiler_ders, columns=degerlendirme_sutunlari, index=range(1, ders_cikti_s + 1))


agirliklar_dict = {degerlendirme_sutunlari[i]: agirliklar[i] for i in range(len(agirliklar))}
df_degerlendirme.loc["Ağırlıklar"] = agirliklar_dict


df_degerlendirme_without_weights = df_degerlendirme.drop(index="Ağırlıklar")  


df_degerlendirme_without_weights["Toplam"] = df_degerlendirme_without_weights.sum(axis=1)


df_degerlendirme["Toplam"] = df_degerlendirme_without_weights["Toplam"]


file_name = "ders_ciktisi_tablosu.xlsx"
if os.path.exists(file_name):
    print(f"{file_name} dosyası zaten mevcut. Yeni dosya oluşturulmadı.")
else:
   
    with pd.ExcelWriter(file_name) as writer:
        df_degerlendirme.to_excel(writer, sheet_name="ders_ciktisi_tablosu", index_label="Ders Çıktısı")

    print(f"{file_name} başarıyla oluşturuldu.")


file_name_ders_ciktisi = "ders_ciktisi_tablosu.xlsx"


if os.path.exists(file_name_ders_ciktisi):
    df_degerlendirme = pd.read_excel(file_name_ders_ciktisi, sheet_name="ders_ciktisi_tablosu", index_col="Ders Çıktısı")
    print(f"{file_name_ders_ciktisi} dosyasından veri okundu.")
else:
    print(f"{file_name_ders_ciktisi} bulunamadı. Lütfen dosyayı oluşturun.")
    exit()


agirliklar = df_degerlendirme.loc["Ağırlıklar"].values
degerlendirme_sutunlari = df_degerlendirme.columns[:-1] 


df_agirlikli_degerlendirme = pd.DataFrame(index=df_degerlendirme.index[:-1], columns=degerlendirme_sutunlari)  


for col in degerlendirme_sutunlari:
    for row in df_degerlendirme.index[:-1]: 
        
        new_value = (df_degerlendirme.loc[row, col] * agirliklar[degerlendirme_sutunlari.tolist().index(col)]) / 100
        df_agirlikli_degerlendirme.loc[row, col] = round(new_value, 2)  


df_agirlikli_degerlendirme["Toplam"] = df_agirlikli_degerlendirme.sum(axis=1)


file_name_agirlikli = "agirlikli_degerlendirme.xlsx"


if os.path.exists(file_name_agirlikli):
    print(f"{file_name_agirlikli} dosyası zaten mevcut. Yeni dosya oluşturulmadı.")
else:

    with pd.ExcelWriter(file_name_agirlikli) as writer:
        df_agirlikli_degerlendirme.to_excel(writer, sheet_name="Ağırlıklı Değerlendirme", index_label="Ders Çıktısı")
    print(f"{file_name_agirlikli} başarıyla oluşturuldu.")







file_name_dersanotlar = "DersANotlar.xlsx"
file_name_agirlikli = "agirlikli_degerlendirme.xlsx"
new_file_name = "agirlikli_ogrenci_notlari.xlsx" 


if os.path.exists(file_name_dersanotlar):
    df_dersanotlar = pd.read_excel(file_name_dersanotlar, sheet_name="Sayfa1", index_col="Ogrenci_No")
    print(f"{file_name_dersanotlar} dosyasından veri okundu.")
else:
    print(f"{file_name_dersanotlar} bulunamadı. Lütfen dosyayı oluşturun.") 

if os.path.exists(file_name_agirlikli):
    df_agirlikli = pd.read_excel(file_name_agirlikli, sheet_name="Ağırlıklı Değerlendirme", index_col="Ders Çıktısı")
    print(f"{file_name_agirlikli} dosyasından veri okundu.")
else:
    print(f"{file_name_agirlikli} bulunamadı. Lütfen dosyayı oluşturun.")
    exit()


with pd.ExcelWriter(new_file_name, engine='openpyxl') as writer: 
    
    start_row = 1

    for student in df_dersanotlar.index:
        student_data = df_dersanotlar.loc[student]

        student_table = pd.DataFrame(columns=df_agirlikli.columns, index=df_agirlikli.index)

        for column in df_agirlikli.columns:
            if column != "Toplam":  
                for row in df_agirlikli.index:
                    student_table.loc[row, column] = round(student_data[column] * df_agirlikli.loc[row, column], 2)


    
        student_table["Toplam"] = student_table.sum(axis=1)

        student_table["Max"] = student_table.apply(lambda row: sum([100 * df_agirlikli.loc[row.name, col] for col in df_agirlikli.columns[:-1]]), axis=1)

        student_table["Başarı Oranı"] = (student_table["Toplam"] / student_table["Max"]) * 100

        
        student_table.to_excel(writer, sheet_name="Tüm Öğrenciler", startcol=0, startrow=start_row, index_label=f"Öğrenci {student}")
        
        
        start_row += len(student_table) + 3  

        
        
    print(f"Tüm öğrencilerin tabloları tek bir sayfada başarıyla oluşturuldu.")



file_name_agirlikli_notlar = "agirlikli_ogrenci_notlari.xlsx"
file_name_ders_program_ciktisi = "ders_program_cikti_iliskisi.xlsx"
new_file_name_2 = "agirlikli_ders_program_ciktisi_notlari.xlsx"

if os.path.exists(file_name_ders_program_ciktisi):
    df_dersanotlar = pd.read_excel(file_name_ders_program_ciktisi, sheet_name="Tablo 1", index_col="Prg Çıktı")
    print(f"{file_name_ders_program_ciktisi} dosyasından veri okundu.")

    df_dersanotlar_sliced = df_dersanotlar.iloc[:, 0:-1]

    column_lists = [df_dersanotlar_sliced[col].tolist() for col in df_dersanotlar_sliced.columns]

    df_dersanotlar_sliced = df_dersanotlar.iloc[:, 0:]

    column_lists1 = [df_dersanotlar_sliced[col].tolist() for col in df_dersanotlar_sliced.columns]

    
else:
    print(f"{file_name_ders_program_ciktisi} bulunamadı. Lütfen dosyayı oluşturun.")

if os.path.exists(file_name_agirlikli_notlar):

    df = pd.read_excel(file_name_agirlikli_notlar, sheet_name="Tüm Öğrenciler")

    student_tables = []
    student_numbers = []

    basarim_orani_listesi = []

    start_row = 0  #
    while start_row < len(df):

        if isinstance(df.iloc[start_row, 0], str) and df.iloc[start_row, 0].startswith("Öğrenci"):

            student_number = df.iloc[start_row, 0]
            student_numbers.append(student_number)

            student_table = df.iloc[start_row + 1:start_row + 6]
            student_tables.append(student_table)

            basarim_orani = student_table.iloc[:, 9].tolist()
            basarim_orani_listesi.append(basarim_orani)

            start_row += len(student_table) + 3
        else:
            start_row += 1

    
    result_lists = []
    for first_item in basarim_orani_listesi:

        result = []
        for second_item in column_lists:
            multiplied_values = [first_value * second_value for first_value, second_value in
                                 zip(first_item, second_item)]
            result.append(multiplied_values)
        result_lists.append(result)
result_lists1=[]
for first_item in basarim_orani_listesi:
        result = []
        for i in range(len(first_item)):  
            second_item = column_lists[i]  
            
            multiplied_values = [first_item[i] * value for value in second_item]
            result.append(multiplied_values)
        result_lists1.append(result)
        
        
        

   
for res in result_lists:
    

    with pd.ExcelWriter(new_file_name_2, engine='openpyxl') as writer:

        df = pd.read_excel("ders_program_cikti_iliskisi.xlsx", sheet_name="Tablo 1")

        start_row = 0
        program_cikti_sayilari = df['Prg Çıktı'].tolist()

        

        for student in student_numbers:
            sheet_name = "Ders Çıktısı"

            student_table = pd.DataFrame({
                'Prg Çıktı': program_cikti_sayilari,

            })

            student_table.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, index=False)

            start_row += len(student_table) + 3


else:
    print()

workbook = load_workbook(new_file_name_2)
worksheet = workbook[sheet_name]

row = 2



prg_cikti_col = 1  

bold_font = Font(bold=True) 
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)  


row_indices = []

for row in range(1, worksheet.max_row + 1):  
    if worksheet.cell(row=row, column=prg_cikti_col).value == "Prg Çıktı":
        row_indices.append(row)
        last_column = len(basarim_orani_listesi[0])+2
        for col in range(1, last_column): 
            cell = worksheet.cell(row=row, column=col)
            cell.font = bold_font  
            cell.border = thin_border  

row_indices2=row_indices

row_indices1=[row - 1 for row in row_indices]

iliski_deg_as=column_lists1[-1]

for idx, row in enumerate(row_indices1):
    
    cell_student = worksheet.cell(row=row, column=1, value=student_numbers[idx])
    cell_student.font = bold_font  
    cell_student.border = thin_border  

    
    cell_output = worksheet.cell(row=row, column=2, value="Ders Çıktısı")
    cell_output.font = bold_font  
    cell_output.border = thin_border  


for idx, row in enumerate(row_indices):
    if idx < len(basarim_orani_listesi):  
        value_list = basarim_orani_listesi[idx]
        
        
        for col_offset, value in enumerate(value_list):
            worksheet.cell(row=row, column=prg_cikti_col + 1 + col_offset, value=value)
        
       
        last_col = prg_cikti_col  + len(value_list)  
        cell = worksheet.cell(row=row, column=last_col + 1) 
        cell.value = "Başarı Oranı"  
        cell.font = bold_font  
        cell.border = thin_border  

row_indices = [row + 1 for row in row_indices]
prg_cikti_col = 2  


for idx, row in enumerate(row_indices):
    column_length = len(column_lists[0])  
    
   
    for col_offset in range(column_length):
        col = 1 + col_offset  
        
        
        for row_num in range(row, row + column_length):  
            
            if col == 1:  
                cell = worksheet.cell(row=row_num, column=col)
                cell.font = bold_font  
                cell.border = thin_border  


for idx, row in enumerate(row_indices):
    if idx < len(result_lists1):  
        value_list = result_lists1[idx]  
        col = prg_cikti_col 
        
        
        
        
        start_row = row  
        
        for sublist in value_list: 
            current_row = start_row 
            for value in sublist:  
                worksheet.cell(row=current_row, column=col, value=value)  
                current_row += 1  
            col += 1  

iliski_deger_icin = []

row_indices2 = [row + 1 for row in row_indices2]
for idx, row in enumerate(row_indices2):
    row_total = []  
    
    
    for i in range(len(column_lists[0])):
        
        total = 0
        
        for offset in range(0,(len(basarim_orani_listesi[0])+1)): 
            right_cell = worksheet.cell(row=row + i, column=prg_cikti_col + offset)
            
            try:
                
                if right_cell.value:
                    total += float(right_cell.value)
            except ValueError:
                continue  
        
        
        row_total.append(total)
    
    
    iliski_deger_icin.append(row_total)
    
agirlikli_basarim_orani = []


for idx, inner_list in enumerate(iliski_deger_icin):
    row_agirlikli = []  
    
    
    for i, value in enumerate(inner_list):
        
        division1 = value / len(basarim_orani_listesi[0]) if len(basarim_orani_listesi[0]) != 0 else 0 
        print(f"division1 (value={value}, len(basarim_orani_listesi[0])={len(basarim_orani_listesi[0])}): {division1}")

        
        division2 = division1 / iliski_deg_as[i] if iliski_deg_as[i] != 0 else 0  
        print(f"division2 (division1={division1}, iliski_deg_as[{i}]={iliski_deg_as[i]}): {division2}")
        
        
        row_agirlikli.append(division2)
    
    
    agirlikli_basarim_orani.append(row_agirlikli)
print(agirlikli_basarim_orani)


print(agirlikli_basarim_orani)
print(iliski_deg_as)
red_bold_font = Font(bold=True, color="FF0000")
row_indices2 = [row + -1 for row in row_indices2]

for idx, row in enumerate(row_indices2):
    
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row, column=col)
        if cell.value == "Başarı Oranı":  
            basarim_orani_col = col  
            break
    else:
        continue  
    
   
    if idx < len(agirlikli_basarim_orani):  
        for i, value in enumerate(agirlikli_basarim_orani[idx]):
            target_cell = worksheet.cell(row=row + 1 + i, column=basarim_orani_col) 
            target_cell.value = value  
            target_cell.font = red_bold_font  
            target_cell.border = thin_border  

workbook.save(new_file_name_2)
